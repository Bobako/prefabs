import openpyxl

wb = openpyxl.load_workbook(filename)
sheet = wb.active  # grab active worksheet
# sheet = wb["some_sheet"] # as option; some_sheet must be in wb.sheetnames

row = 1
while sheet.cell(row=row, column=1).value:  # pass over values
    do_some_fucking_magic()
    row += 1

wb.close()
# wb.save(filename)
